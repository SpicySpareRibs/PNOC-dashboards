from datetime import timedelta
from typing import Protocol
from recompiler import Datafile,Recompiler
import pandas
from pandas import Timestamp
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np
import warnings

warnings.simplefilter(action='ignore', category=FutureWarning)
class ProcMonitoringCompiler():
    def __init__(self,datafile):
        self._datafile = datafile
        self._dataframe = None

        
        self._header_constraints = None
        self._stages = None
        self._record_stage = None
        self._record = None
        self._record_codes = None

    def get_headers(self):
        pass
    def set_constraints(self):
        pass
    def get_hidden_cols(self,max):
        workbook = openpyxl.load_workbook(self._datafile.filepath,data_only=True)
        ws = workbook["RMB-PMR 2024 "]
        
        max_col = ws.max_column
        cols = [get_column_letter(i) for i in range(1, max_col+1)]

        hidden_cols = []
        last_hidden = 0
        for i, col in enumerate(cols):
            if i >= max:
                break
        # Column is hidden
            if ws.column_dimensions[col].hidden:
                hidden_cols.append(i)
                # Last column in the hidden group
                last_hidden = ws.column_dimensions[col].max
        # Appending column if more columns in the group
            elif i+1 <= last_hidden:
                hidden_cols.append(i)
            
        
        return hidden_cols 
    def replace_header(self,level0,level1):
            pass

    def drop_empty_in_col(self,dataframe,col_index):
        df = dataframe
        return df.dropna(subset=[df.columns[0]])
        
        

        
    def replace_date(self, timestamp):
        
        try: 
            
            
            return timestamp.date() 
        except:
            print("Error: input is not a timestamp")
            return timestamp
        
    def cascade_headers(self,df: pandas.DataFrame,top_levels,bottom_level):# -> list[Any]:
        
        bottom_row = df.columns.get_level_values(bottom_level).to_list()
        
        
        items_per_level = dict((l,df.columns.get_level_values(l).to_list()) for l in top_levels)
        for j in range(0,len(bottom_row)):
                if "Unnamed" in bottom_row[j]:
                    for i in reversed(top_levels):
                        if "Unnamed" not in items_per_level[i][j]:
                            bottom_row[j] = items_per_level[i][j]
                            break
                else:
                    bottom_row[j] = items_per_level[0][j] + bottom_row[j]
        final_headers = bottom_row
        return final_headers
    
    def read_datafile(self):
        df = pandas.read_excel(self._datafile.filepath,sheet_name="RMB-PMR 2024 ",header=[6,7])
        
        hidden_cols = self.get_hidden_cols(len(df.columns))
        
        all_levels = self.cascade_headers(df,[0],1)
        
        df.columns = df.columns.droplevel(0)
        df.columns = all_levels
        
        df.drop(df.columns[hidden_cols],axis=1,inplace=True)

        
    

        df = self.drop_empty_in_col(df,0)
        
        df["number"] = df.index
        self._dataframe = df

    def rename_columns(self, df:pandas.DataFrame):
        new_cols = ["Codes",
            "Description",
            "End-User",
            "Is Early Procurement",
            "Mode of Procurement",
            "Pre-Proc Conference",
            "Ads/Post of IAEB",
            "Sub/Open of Bids",
            "Bid Evaluation",
            "Post Qual",
            "Notice of Award",
            "Contract Signing",
            "Notice to Proceed",
            "Delivery/Completion",
            "Source of Funds",
            "Total-Alloted",
            "Alloted Budget-MOOE",
            "Alloted Budget-CO",
            "Total-Contract",
            "Contract Cost-MOOE",
            "Contract Cost-CO",
            "Remarks",
            "Number"]
        for row,col in enumerate(df.columns):
            print(row,col,"-",new_cols[row],"\n")

        
        df.columns= new_cols
        
        return df
    def create_stage_rank_table(self):
        self._stage_rank = pandas.DataFrame(columns=["stage","rank"])
        stages = ["Pre-Proc Conference",
            "Ads/Post of IAEB",
            "Sub/Open of Bids",
            "Bid Evaluation",
            "Post Qual",
            "Notice of Award",
            "Contract Signing",
            "Notice to Proceed",
            "Delivery/Completion",
            ]
        for rank, stage in enumerate(stages):
            self._stage_rank.loc[rank,self._stage_rank.columns] = rank,stage
    def reformat(self):
        df = self._dataframe
        
        df.drop(df.columns[[7,8,12,24,25,26,27,28,29,30]],axis=1,inplace=True)
        
        df = self.rename_columns(df)
        print(df['Sub/Open of Bids'])

        self.create_record_codes_table(df)
        self.create_monitoring_record_table(df)
        self.create_stage_rank_table()
        self.create_record_stage_table(df)
        self._dataframe = df
    def split_codes(self, code_list: str):
        print(code_list)
        return code_list.split("/")
    
    def create_record_codes_table(self,df:pandas.DataFrame):
        self._record_codes = pandas.DataFrame(columns=["record number","code"])
        
        for row_number,row in df[["Number","Codes"]].iterrows():
            
            for code in self.split_codes(row[1]):
                self._record_codes.loc[len(self._record_codes), self._record_codes.columns] = row[0],code
        

    def create_record_stage_table(self,df:pandas.DataFrame):
        self._record_stage = pandas.DataFrame(columns=["Record Number","Stage Rank","Date"])


        recorded_stages = 9

        for row_number, row in df[[
            "Number",
            "Pre-Proc Conference",
            "Ads/Post of IAEB",
            "Sub/Open of Bids",
            "Bid Evaluation",
            "Post Qual",
            "Notice of Award",
            "Contract Signing",
            "Notice to Proceed",
            "Delivery/Completion",]].iterrows():
                is_NA_row = row.isna()
                
                last_recorded_date = None
                date = None
                for stage_rank in range(1,recorded_stages+1):

                    if not is_NA_row[stage_rank]:
                        
                        if not isinstance(row[stage_rank],Timestamp) and isinstance(row[stage_rank],str):
                            date = last_recorded_date + timedelta(days=int(row[stage_rank].replace('days','')))
                            print("last ",last_recorded_date,"+ ",row[stage_rank].replace('days',''),date)
                            last_recorded_date = date
                            
                        else:
                            last_recorded_date = row[stage_rank]
                            date = row[stage_rank]
                        
                        self._record_stage.loc[len(self._record_stage),self._record_stage.columns] = row[0],stage_rank,date
                    
    def create_monitoring_record_table(self,df:pandas.DataFrame):
        self._monitoring_record = pandas.DataFrame(columns=["Record Number", 
                                                            "Description",
                                                            "End user",
                                                            "Is Early Procurement",
                                                            "Mode Of Procurement", 
                                                            "Source Of Funds",
                                                            "Alloted Budget",
                                                            "Contract Cost",
                                                            "Type",
                                                            "Remarks"])
        
        for row_number,row in df[["Number",
                        "Description",
                       "End-User",
                       "Is Early Procurement",
                       "Mode of Procurement",
                        "Source of Funds",
                        "Total-Alloted",
                        "Alloted Budget-MOOE",
                        "Alloted Budget-CO",
                        "Total-Contract", 
                        "Contract Cost-MOOE",
                        "Contract Cost-CO",
                        "Remarks"]].iterrows() :
            
            
            is_NA_row = row.isna()
            if is_NA_row[8]:
                contract_type = "MOOE"
                     
            else:
                contract_type = "CO"
                
            self._monitoring_record.loc[len(self._monitoring_record),self._monitoring_record.columns] = row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[9], contract_type,row[12]
    def export(self):
        with pandas.ExcelWriter('PMR.xlsx',datetime_format="MM-DD-YYYY",date_format="MM-DD-YYYY") as writer:     
            self._dataframe.to_excel(writer,sheet_name="original",index=False)
            self._record_codes.to_excel(writer, sheet_name="record codes",index=False)
            self._monitoring_record.to_excel(writer, sheet_name="monitoring record",index=False)
            self._stage_rank.to_excel(writer,sheet_name="stage rank",index=False)
            self._record_stage.to_excel(writer,sheet_name="record stage",index=False)

# Under Dev            
df = Datafile(filepath="Procurement\\PMR.xlsx",filetype="Procurement Monitoring Report")
test = ProcMonitoringCompiler(datafile=df)
test.read_datafile()
test.reformat()
test.export()
